"""ATLAS v2 — Telegram-to-Claude personal agent for Dr. Madhav Chowdhry.

Telegram = interface · Claude = brain · Drive/Sheets = filing cabinet
Gmail/Calendar = live feeds (polled every few minutes, triaged silently)

Protocol-timed nudges (TIMEZONE):
  05:15  gym confirmation (knows upper/lower day)
  07:30  breakfast photo + morning brief (+ calendar; Sunday = weekly review)
  13:00  lunch photo
  17:00  snack check
  20:30  dinner photo
  21:30  evening close
  22:00  skincare + lights-out warning
Live jobs:
  every 5 min   Gmail poll -> Claude triage -> interrupt ONLY if it matters
  every 15 min  Calendar look-ahead -> event reminders

Commands (type in Telegram):
  papers / nuos / hospital / cases / health  -> that tracker's status
  add to papers: <anything>                  -> Claude files it as a row
  social                                     -> post/reply queue for approval
  SEND                                       -> send the pending email draft
"""
import base64
import io
import json
import logging
import os
import re
from datetime import datetime, time as dtime
from zoneinfo import ZoneInfo

import anthropic
from dotenv import load_dotenv
from PIL import Image
from telegram import Update
from telegram.ext import Application, ContextTypes, MessageHandler, filters

from gdrive import DriveStore
from gmail_cal import GoogleLive

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("atlas")

BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
USER_ID = int(os.environ["TELEGRAM_USER_ID"])
MODEL = os.getenv("ATLAS_MODEL", "claude-sonnet-4-6")
TZ = ZoneInfo(os.getenv("TIMEZONE", "Asia/Kolkata"))

MEMORY_FILE = "memory.json"
STATE_FILE = "state.json"
PENDING_EMAIL = "pending_email.json"
MEMORY_TURNS = 40

# --- zero-config bootstrap for cloud deploys (Railway etc.) ---
# 1) service_account.json can arrive as an env var instead of a file
_sa_env = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
_sa_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "service_account.json")
if _sa_env and not os.path.exists(_sa_path):
    with open(_sa_path, "w", encoding="utf-8") as _f:
        _f.write(_sa_env)

# 2) first boot: build the Drive tree automatically if DRIVE_FOLDER_ID is set
if not os.path.exists("atlas_ids.json") and os.getenv("DRIVE_FOLDER_ID"):
    import subprocess
    import sys as _sys
    subprocess.run(
        [_sys.executable, "setup_drive.py", os.environ["DRIVE_FOLDER_ID"]],
        check=True,
    )

claude = anthropic.Anthropic()
store = DriveStore(_sa_path)
glive = GoogleLive()

TRACKERS = {"papers": "papers", "nuos": "nuos", "hospital": "hospital",
            "cases": "case_log", "health": "health_log"}


# ---------------------------------------------------------------- small stores

def _load(path, default):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def _save(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)


def load_memory():
    return _load(MEMORY_FILE, [])


def remember(role, text):
    mem = load_memory()
    mem.append({"role": role, "content": text})
    _save(MEMORY_FILE, mem[-MEMORY_TURNS:])


# ---------------------------------------------------------------- system prompt

def load_system() -> str:
    with open("atlas_persona.txt", encoding="utf-8") as f:
        persona = f.read()
    with open("atlas_context.md", encoding="utf-8") as f:
        context = f.read()
    now = datetime.now(TZ)
    pending = _load(PENDING_EMAIL, None)
    pending_note = ""
    if pending:
        pending_note = (
            "\n\nPENDING EMAIL DRAFT (awaiting his approval; he sends by typing "
            "SEND): " + json.dumps(pending)
            + "\nIf he asks for edits, include the full revised draft inside "
            '<draft_json>{"to":"","subject":"","body":""}</draft_json> tags at '
            "the end of your reply (the pipeline stores it and strips the tags)."
        )
    return (
        persona
        + "\n\nCURRENT DATE/TIME: " + now.strftime("%A, %d %B %Y, %H:%M (%Z)")
        + pending_note
        + "\n\n=== MASTER CONTEXT ===\n" + context
    )


def normalized(msgs):
    out = []
    for m in msgs:
        if (out and out[-1]["role"] == m["role"]
                and isinstance(out[-1]["content"], str)
                and isinstance(m["content"], str)):
            out[-1]["content"] += "\n" + m["content"]
        else:
            out.append({"role": m["role"], "content": m["content"]})
    if out and out[0]["role"] != "user":
        out.insert(0, {"role": "user", "content": "(context resumes)"})
    return out


def ask_claude(messages, max_tokens=1200):
    resp = claude.messages.create(model=MODEL, max_tokens=max_tokens,
                                  system=load_system(),
                                  messages=normalized(messages))
    return "".join(b.text for b in resp.content if b.type == "text").strip()


def parse_json_reply(text):
    cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip(), flags=re.M)
    start, end = cleaned.find("{"), cleaned.rfind("}")
    return json.loads(cleaned[start:end + 1])


DRAFT_RE = re.compile(r"<draft_json>\s*(\{.*?\})\s*</draft_json>", re.S)


def extract_draft(reply: str) -> str:
    """If Claude embedded a revised email draft, store it and strip the tags."""
    m = DRAFT_RE.search(reply)
    if m:
        try:
            _save(PENDING_EMAIL, json.loads(m.group(1)))
        except json.JSONDecodeError:
            log.warning("Draft tags present but JSON invalid")
        reply = DRAFT_RE.sub("", reply).strip()
    return reply


# ---------------------------------------------------------------- text messages

WEIGHT_RE = re.compile(r"\bweight\b\s*[:\-]?\s*(\d{2,3}(?:\.\d+)?)", re.I)
ADD_RE = re.compile(r"^add to (papers|nuos|hospital)\s*[:\-]\s*(.+)$", re.I | re.S)


async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user is None or update.effective_user.id != USER_ID:
        return
    text = (update.message.text or "").strip()
    low = text.lower()

    # --- SEND: approved email goes out (prime rule 1: explicit, in the moment)
    if low == "send":
        pending = _load(PENDING_EMAIL, None)
        if not pending:
            await update.message.reply_text("No email is staged for sending.")
            return
        if not glive.ok:
            await update.message.reply_text(
                "Gmail isn't connected on the server (token.json missing).")
            return
        try:
            glive.send_email(pending["to"], pending["subject"], pending["body"])
            os.remove(PENDING_EMAIL)
        except Exception:
            log.exception("Send failed")
            await update.message.reply_text("Send failed — check server log.")
            return
        remember("user", "SEND (approved the staged email)")
        remember("assistant", f"Sent email to {pending['to']}: {pending['subject']}")
        await update.message.reply_text(
            f"Sent — {pending['to']} · {pending['subject']}")
        return

    # --- tracker status commands
    if low in TRACKERS:
        try:
            rows = store.tracker_rows(TRACKERS[low])
        except Exception:
            log.exception("Tracker read failed")
            await update.message.reply_text("Couldn't read that tracker just now.")
            return
        prompt = (f"[command: {low}] Current rows of the {low} tracker:\n"
                  + json.dumps(rows)
                  + "\nGive a tight status: what's moving, what's stalled, the "
                  "one next action. Telegram-short.")
        reply = ask_claude(load_memory() + [{"role": "user", "content": prompt}])
        remember("user", text)
        remember("assistant", reply)
        await update.message.reply_text(reply)
        return

    # --- add to a tracker
    m = ADD_RE.match(text)
    if m:
        key, item = m.group(1).lower(), m.group(2).strip()
        headers = store.tracker_rows(TRACKERS[key], n=1)[0]
        prompt = (f'He wants to add to the {key} tracker: "{item}"\n'
                  f"Sheet columns: {json.dumps(headers)}\n"
                  'Respond ONLY with minified JSON: {"row": [...one value per '
                  'column, today\'s date for Updated...], "confirm": "one-line '
                  'confirmation in ATLAS voice"}')
        try:
            meta = parse_json_reply(ask_claude(
                load_memory() + [{"role": "user", "content": prompt}],
                max_tokens=500))
            store.tracker_append(TRACKERS[key], meta["row"])
        except Exception:
            log.exception("Tracker append failed")
            await update.message.reply_text("Couldn't file that — try rephrasing.")
            return
        remember("user", text)
        remember("assistant", meta["confirm"])
        await update.message.reply_text(meta["confirm"])
        return

    # --- social queue (pull, never push)
    if low == "social":
        prompt = ("[command: social] He's free and pulling the social queue. "
                  "From recent memory and the master context (Domain 7: one "
                  "LinkedIn teaching post/week from anonymised case or paper), "
                  "produce: 1) one drafted LinkedIn post ready to approve, "
                  "2) anything else pending. He approves/edits here; posting "
                  "is manual or via the Claude.ai session. Keep it tight.")
        reply = ask_claude(load_memory() + [{"role": "user", "content": prompt}])
        remember("user", text)
        remember("assistant", reply)
        await update.message.reply_text(reply)
        return

    # --- weight logging
    pipeline_note = ""
    w = WEIGHT_RE.search(text)
    if w:
        now = datetime.now(TZ)
        try:
            store.append_health([now.strftime("%Y-%m-%d"), now.strftime("%H:%M"),
                                 "weight", w.group(1) + " kg", ""])
            pipeline_note = ("\n\n[system note: weight " + w.group(1)
                             + " kg logged to the Health Log — confirm briefly]")
        except Exception:
            log.exception("Health log append failed")
            pipeline_note = ("\n\n[system note: weight logging FAILED — tell "
                             "him honestly]")

    # --- normal conversation
    try:
        reply = ask_claude(load_memory()
                           + [{"role": "user", "content": text + pipeline_note}])
    except Exception:
        log.exception("Claude call failed")
        await update.message.reply_text(
            "ATLAS hit an API error. Try again in a minute.")
        return
    reply = extract_draft(reply)
    remember("user", text)
    remember("assistant", reply)
    await update.message.reply_text(reply)


# ---------------------------------------------------------------- photos

def autocrop(img, threshold=24):
    gray = img.convert("L")
    mask = gray.point(lambda p: 255 if p > threshold else 0)
    bbox = mask.getbbox()
    return img.crop(bbox) if bbox else img


def photo_prompt(caption):
    return (
        'A photo has arrived in the ATLAS Telegram chat. Caption (may be empty): "'
        + caption + '"\n\n'
        "Classify and extract. Respond ONLY with minified JSON, no markdown "
        "fences, exactly these keys:\n"
        '{"type": "xray|scan|meal|other", "identifiers_visible": false, '
        '"region": "", "view": "", "findings": "", '
        '"suggested_filename": "snake_case_no_extension", '
        '"log_note": "", "reply": ""}\n\n'
        "Rules:\n"
        "- identifiers_visible = true if ANY patient name, MRN, hospital number, "
        "DOB or other identifying text is burned into a medical image. Inspect "
        "corners and edges carefully.\n"
        "- xray/scan: region, view, findings = one-line cold read "
        "(fracture pattern/classification if evident, hardware).\n"
        "- meal: log_note = short description vs his diet template with rough "
        "protein estimate.\n"
        "- reply = short Telegram message: one-line filing confirmation + one "
        "teaching pearl (medical) or one diet note (meal). ATLAS voice."
    )


def safe_name(name):
    name = re.sub(r"[^a-zA-Z0-9_-]+", "_", name).strip("_").lower()
    return name or "image"


async def on_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user is None or update.effective_user.id != USER_ID:
        return
    caption = update.message.caption or ""
    tg_file = await update.message.photo[-1].get_file()
    raw = bytes(await tg_file.download_as_bytearray())
    img = Image.open(io.BytesIO(raw)).convert("RGB")
    img = autocrop(img)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=90)
    data = buf.getvalue()
    b64 = base64.b64encode(data).decode()

    content = [
        {"type": "image",
         "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
        {"type": "text", "text": photo_prompt(caption)},
    ]
    try:
        meta = parse_json_reply(ask_claude(
            load_memory() + [{"role": "user", "content": content}],
            max_tokens=800))
    except Exception:
        log.exception("Photo analysis failed")
        await update.message.reply_text(
            "Couldn't analyse that image just now. Resend in a minute?")
        return

    kind = meta.get("type", "other")
    now = datetime.now(TZ)
    fname = (now.strftime("%Y-%m-%d_%H%M_")
             + safe_name(meta.get("suggested_filename", "image")) + ".jpg")

    if kind in ("xray", "scan") and meta.get("identifiers_visible"):
        remember("user", f"[photo NOT saved — identifiers visible] {caption}")
        remember("assistant", "Asked him to crop identifiers and resend.")
        await update.message.reply_text(
            "Not filed — patient identifiers are visible. Crop the name/MRN and "
            "resend; the clean path applies to storage too.")
        return

    try:
        if kind in ("xray", "scan"):
            link = store.upload_image("cases", fname, data)
            store.append_case([now.strftime("%Y-%m-%d"), now.strftime("%H:%M"),
                               kind, meta.get("region", ""), meta.get("view", ""),
                               meta.get("findings", ""), "", "", caption, link])
        elif kind == "meal":
            link = store.upload_image("health", fname, data)
            store.append_health([now.strftime("%Y-%m-%d"), now.strftime("%H:%M"),
                                 "meal", meta.get("log_note", caption), link])
        else:
            store.upload_image("health", fname, data)
    except Exception:
        log.exception("Filing failed")
        await update.message.reply_text(
            "Image analysed but Drive filing failed — check server log.")
        return

    reply = meta.get("reply") or f"Filed as {fname}."
    remember("user", f"[photo filed: {fname}] {caption}")
    remember("assistant", reply)
    await update.message.reply_text(reply)


# ---------------------------------------------------------------- documents (DocPulse reports)

DOC_EXTS = {".xlsx": "excel", ".xls": "excel", ".csv": "csv", ".pdf": "pdf"}


def table_text(data: bytes, kind: str) -> str:
    """Flatten a report into text Claude can read. Caps size defensively."""
    if kind == "csv":
        return data.decode("utf-8", errors="replace")[:60000]
    if kind == "excel":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"### Sheet: {ws.title}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 400:
                    lines.append("...(truncated)")
                    break
                lines.append(",".join("" if c is None else str(c) for c in row))
        return "\n".join(lines)[:60000]
    return ""


def report_prompt(filename: str, caption: str, body: str) -> str:
    return (
        f'A hospital report file arrived: "{filename}". Caption: "{caption}"\n'
        "It is (very likely) a DocPulse export from the family hospital. "
        "Contents follow.\n\n" + body + "\n\n"
        "Analyse per Domain 4: reconcile the three flows where visible "
        "(clinical events vs stock vs money), spot gaps or anomalies "
        "(unbilled events, concessions without names, outliers).\n"
        "Respond ONLY with minified JSON:\n"
        '{"report_type": "", "daily_row": ["YYYY-MM-DD","report","OPD","IPD",'
        '"collections","pharmacy","claims","flags",""] , "flags": [], '
        '"reply": ""}\n'
        "- daily_row matches the Hospital Daily sheet columns; leave unknown "
        "cells as empty strings, use the report's own date.\n"
        "- flags: each a one-line anomaly worth his attention (empty list if "
        "clean).\n"
        "- reply: short Telegram summary in ATLAS voice — the numbers that "
        "matter + flags. No lecture."
    )


async def on_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user is None or update.effective_user.id != USER_ID:
        return
    doc = update.message.document
    name = doc.file_name or "report"
    ext = os.path.splitext(name)[1].lower()
    if ext not in DOC_EXTS:
        await update.message.reply_text(
            f"I can ingest .xlsx/.xls/.csv/.pdf reports; {ext or 'that'} "
            "isn't one of them.")
        return
    caption = update.message.caption or ""
    tg_file = await doc.get_file()
    data = bytes(await tg_file.download_as_bytearray())

    kind = DOC_EXTS[ext]
    if kind == "pdf":
        content = [
            {"type": "document",
             "source": {"type": "base64", "media_type": "application/pdf",
                        "data": base64.b64encode(data).decode()}},
            {"type": "text", "text": report_prompt(name, caption, "(see PDF)")},
        ]
    else:
        try:
            body = table_text(data, kind)
        except Exception:
            log.exception("Report parse failed")
            await update.message.reply_text(
                "Couldn't open that file — is it a standard export?")
            return
        content = report_prompt(name, caption, body)

    try:
        meta = parse_json_reply(ask_claude(
            load_memory() + [{"role": "user", "content": content}],
            max_tokens=1000))
    except Exception:
        log.exception("Report analysis failed")
        await update.message.reply_text(
            "Report received but analysis failed — resend or check the format.")
        return

    now = datetime.now(TZ)
    stamped = now.strftime("%Y-%m-%d_") + safe_name(os.path.splitext(name)[0]) + ext
    mimes = {"excel": "application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet",
             "csv": "text/csv", "pdf": "application/pdf"}
    try:
        store.upload_file("reports", stamped, data, mimes[kind])
        if meta.get("daily_row"):
            store.tracker_append("hospital_daily", meta["daily_row"])
    except Exception:
        log.exception("Report filing failed")
        await update.message.reply_text(
            "Analysed, but Drive filing failed — check server log.")
        return

    reply = meta.get("reply") or f"Report filed: {stamped}"
    remember("user", f"[hospital report filed: {stamped}] {caption}")
    remember("assistant", reply)
    await update.message.reply_text(reply)


# ---------------------------------------------------------------- nudges

async def push(context, prompt, max_tokens=1200):
    try:
        reply = ask_claude(load_memory() + [{"role": "user", "content": prompt}],
                           max_tokens=max_tokens)
    except Exception:
        log.exception("Scheduled push failed")
        return
    reply = extract_draft(reply)
    remember("assistant", reply)
    await context.bot.send_message(chat_id=USER_ID, text=reply)


async def job_gym(context):
    await push(context, "[scheduled trigger, 05:15] Gym nudge. Work out from "
               "today's weekday which session it is (Mon/Thu lower, Tue/Fri "
               "upper, else rest/walk) and nudge accordingly. 2 sentences max.")


async def job_morning(context):
    cal = ""
    if glive.ok:
        try:
            cal = "\nToday's calendar: " + json.dumps(glive.today(TZ))
        except Exception:
            log.exception("Calendar read failed")
    extra = ""
    if datetime.now(TZ).weekday() == 6:
        rows = []
        try:
            rows = store.recent_health(15)
        except Exception:
            pass
        extra = (" Sunday: fold in the weekly review — weight trend from these "
                 "Health Log rows, QBank/radiology check, one paper lane, "
                 "visibility rep, next week's one priority per active domain. "
                 "Rows: " + json.dumps(rows))
    await push(context, "[scheduled trigger, 07:30] Morning brief per RITUAL "
               "PROTOCOL: greeting + how is he, breakfast photo ask, ONE "
               "priority, countdowns from KEY DATES, domain flags, approvals "
               "pending. Under 10 minutes' reading." + cal + extra)


async def job_lunch(context):
    await push(context, "[scheduled trigger, 13:00] Lunch photo ask + one "
               "mid-day check tied to today's priority. Max 3 sentences.")


async def job_snack(context):
    await push(context, "[scheduled trigger, 17:00] Snack check (protein, no "
               "fried) + quick pulse on the day's ONE priority. Max 3 sentences.")


async def job_dinner(context):
    await push(context, "[scheduled trigger, 20:30] Dinner photo ask. "
               "2 sentences max.")


async def job_close(context):
    await push(context, "[scheduled trigger, 21:30] Evening close per RITUAL "
               "PROTOCOL: cases logged, meals photographed, one line moved/stuck.")


async def job_night(context):
    await push(context, "[scheduled trigger, 22:00] Skincare nudge (PM routine; "
               "track the adapalene/tret ramp from context) + lights-out at "
               "22:30 for the 5:30 gym. 2 sentences max.")


# ---------------------------------------------------------------- live feeds

async def job_email_poll(context):
    if not glive.ok:
        return
    state = _load(STATE_FILE, {"triaged": [], "notified_events": []})
    try:
        msgs = glive.unread_messages()
    except Exception:
        log.exception("Gmail poll failed")
        return
    new = [m for m in msgs if m["id"] not in state["triaged"]]
    if not new:
        return
    state["triaged"] = (state["triaged"] + [m["id"] for m in new])[-300:]
    _save(STATE_FILE, state)

    prompt = ("[email triage — silent unless it matters] New unread emails:\n"
              + json.dumps(new)
              + '\nRespond ONLY with minified JSON: {"notify": false, '
              '"message": "", "draft": null}\n'
              "- notify=true ONLY for genuinely important mail (collaborators "
              "like McPherson/Oxford/AMU/fellowship, deadlines, NuOs, family "
              "hospital, anything urgent). Newsletters/promos/receipts: "
              "notify=false, empty message.\n"
              "- message: the short Telegram alert (sender + gist + your "
              "recommendation).\n"
              '- draft: {"to","subject","body"} reply draft if one is '
              "warranted, else null. If draft present, end message with: "
              "'Reply SEND to send, or tell me the edits.'")
    try:
        meta = parse_json_reply(ask_claude(
            load_memory() + [{"role": "user", "content": prompt}],
            max_tokens=900))
    except Exception:
        log.exception("Triage failed")
        return
    if meta.get("notify") and meta.get("message"):
        if meta.get("draft"):
            _save(PENDING_EMAIL, meta["draft"])
        remember("assistant", meta["message"])
        await context.bot.send_message(chat_id=USER_ID, text=meta["message"])


async def job_calendar_poll(context):
    if not glive.ok:
        return
    state = _load(STATE_FILE, {"triaged": [], "notified_events": []})
    try:
        events = glive.next_half_hour(TZ)
    except Exception:
        log.exception("Calendar poll failed")
        return
    fresh = [e for e in events
             if e["start"] + e["title"] not in state["notified_events"]]
    if not fresh:
        return
    state["notified_events"] = (state["notified_events"]
                                + [e["start"] + e["title"] for e in fresh])[-100:]
    _save(STATE_FILE, state)
    lines = "\n".join(f"{e['start']} — {e['title']}" for e in fresh)
    msg = "Upcoming within 30 min:\n" + lines
    remember("assistant", msg)
    await context.bot.send_message(chat_id=USER_ID, text=msg)


# ---------------------------------------------------------------- main

def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(MessageHandler(filters.PHOTO, on_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, on_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))

    jq = app.job_queue
    jq.run_daily(job_gym, dtime(5, 15, tzinfo=TZ))
    jq.run_daily(job_morning, dtime(7, 30, tzinfo=TZ))
    jq.run_daily(job_lunch, dtime(13, 0, tzinfo=TZ))
    jq.run_daily(job_snack, dtime(17, 0, tzinfo=TZ))
    jq.run_daily(job_dinner, dtime(20, 30, tzinfo=TZ))
    jq.run_daily(job_close, dtime(21, 30, tzinfo=TZ))
    jq.run_daily(job_night, dtime(22, 0, tzinfo=TZ))
    jq.run_repeating(job_email_poll, interval=300, first=30)
    jq.run_repeating(job_calendar_poll, interval=900, first=60)

    log.info("ATLAS v2 online. TZ %s, model %s, Gmail/Cal: %s",
             TZ, MODEL, "connected" if glive.ok else "NOT connected")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
